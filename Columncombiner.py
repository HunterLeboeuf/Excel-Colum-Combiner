import pandas as pd
from openpyxl import load_workbook

# Change the file path to the Excel file you are working with 
df = pd.read_excel("C:\\Users\\Hunter\\Desktop\\test.xlsx")


# List's that we will set each of our columns too 
list1 = [ ]
list2 = [ ]
list3 = [ ]

# Gather the data using pandas, change t1 and t2 to what ever the title of the column is  
list1 = df.loc[:, "t1"]
list2 = df.loc[:, "t2"]

# Add both column 1 and column 2 for column 3 using extend 
list3.extend(list1)
list3.extend(list2)

# Loads the workbook to print list 3 to change to the Excel you are working in 
wb = load_workbook('C:\\Users\\Hunter\\Desktop\\test.xlsx')
ws = wb.active

# row incrementer set below title of row
v = 2
# Loop through list3 and add it to to specified location V is to increment the row change column to location needed to print the comined list 
for x in list3:
    ws.cell(column=3, row=v, value=x)
    v = v + 1
    
# Save the workbook change to your .xlsx file path     
wb.save('C:\\Users\\Hunter\\Desktop\\test.xlsx')
